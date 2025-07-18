# Para subir archivo tipo foto al servidor
from werkzeug.utils import secure_filename
import uuid  # Modulo de python para crear un string

from conexion.conexionBD import connectionBD  # Conexión a BD

import datetime
import re
import os

from os import remove  # Modulo  para remover archivo
from os import path  # Modulo para obtener la ruta o directorio


import openpyxl  # Para generar el excel
# biblioteca o modulo send_file para forzar la descarga
from flask import send_file, session


#Se agrega estas lineas de codigo adicionales del programa
import firebase_admin
from firebase_admin import db # Necesitamos 'db' para interactuar con Realtime Database
import io
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image as ExcelImage # Para incrustar imágenes en Excel
import logging
from openpyxl import load_workbook

logger = logging.getLogger(__name__)
#Hasta aqui 


def accesosReporte():
    """
    Reporte de accesos usando el nuevo esquema (tabla acceso y usuario).
    Si es admin, ve todos los accesos; si no, solo los de su cédula.
    """
    if session['rol'] == 1:  # Administrador
        try:
            with connectionBD() as conexion_MYSQLdb:
                with conexion_MYSQLdb.cursor(dictionary=True) as cursor:
                    querySQL = ("""
                        SELECT 
                            a.id_acceso, 
                            u.cedula, 
                            a.fecha_acceso AS fecha, 
                            a.tipo_acceso, 
                            a.ip_address,
                            r.nombre_rol AS rol
                        FROM acceso a 
                        LEFT JOIN usuario u ON u.id_usuario = a.id_usuario
                        LEFT JOIN roles r ON u.id_rol = r.id_rol
                        ORDER BY a.fecha_acceso DESC
                    """)
                    cursor.execute(querySQL)
                    accesosBD = cursor.fetchall()
                return accesosBD
        except Exception as e:
            print(f"Error en la función accesosReporte para admin: {e}")
            return None
    else:
        try:
            cedula = session['cedula']
            with connectionBD() as conexion_MYSQLdb:
                with conexion_MYSQLdb.cursor(dictionary=True) as cursor:
                    querySQL = ("""
                        SELECT 
                            a.id_acceso, 
                            u.cedula, 
                            a.fecha_acceso AS fecha, 
                            a.tipo_acceso, 
                            a.ip_address,
                            r.nombre_rol AS rol
                        FROM acceso a 
                        LEFT JOIN usuario u ON u.id_usuario = a.id_usuario
                        LEFT JOIN roles r ON u.id_rol = r.id_rol
                        WHERE u.cedula = %s
                        ORDER BY a.fecha_acceso DESC
                    """)
                    cursor.execute(querySQL, (cedula,))
                    accesosBD = cursor.fetchall()
                return accesosBD
        except Exception as e:
            print(f"Error en la función accesosReporte para usuario: {e}")
            return None


def generarReporteExcel():
    dataAccesos = accesosReporte()
    wb = openpyxl.Workbook()
    hoja = wb.active

    # Encabezado solo con columnas reales
    cabeceraExcel = ("ID ACCESO", "CÉDULA", "FECHA/HORA", "TIPO ACCESO", "ROL")
    hoja.append(cabeceraExcel)

    # Si dataAccesos es None, no iterar
    if dataAccesos:
        for registro in dataAccesos:
            id_acceso = registro.get('id_acceso')
            cedula = registro.get('cedula')
            fecha = registro.get('fecha')
            tipo_acceso = registro.get('tipo_acceso')
            rol = registro.get('rol')
            hoja.append((id_acceso, cedula, fecha, tipo_acceso, rol))

    fecha_actual = datetime.datetime.now()
    cedula_sesion = session.get('cedula', 'anonimo') 
    archivoExcel = f"Reporte_accesos_{cedula_sesion}_{fecha_actual.strftime('%Y_%m_%d_%H%M%S')}.xlsx"
    carpeta_descarga = "static/downloads-excel"
    ruta_descarga = os.path.join(os.path.dirname(
        os.path.abspath(__file__)), carpeta_descarga)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
        # Dando permisos a la carpeta
        os.chmod(ruta_descarga, 0o755)

    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    # Enviar el archivo como respuesta HTTP
    return send_file(ruta_archivo, as_attachment=True)


def buscarAreaBD(search):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            a.id_area,
                            a.nombre_area
                        FROM area AS a
                        WHERE a.nombre_area LIKE %s 
                        ORDER BY a.id_area DESC
                    """)
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda

    except Exception as e:
        print(f"Ocurrió un error en def buscarEmpleadoBD: {e}")
        return []


# Lista de Usuarios creados
def lista_usuariosBD():
    try:
        with connectionBD() as conexion_MYSQLdb:
            with conexion_MYSQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                    SELECT id_usuario, cedula, nombre, apellido, email, telefono, fecha_creacion, activo, id_rol,
                        especialidad, anos_experiencia, certificaciones
                    FROM usuario
                """
                cursor.execute(querySQL)
                usuariosBD = cursor.fetchall()
        return usuariosBD
    except Exception as e:
        print(f"Error en lista_usuariosBD : {e}")
        return []

def lista_areasBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT id_area, nombre_area, descripcion FROM area"
                cursor.execute(querySQL,)
                areasBD = cursor.fetchall()
        return areasBD
    except Exception as e:
        print(f"Error en lista_areas : {e}")
        return []

# Eliminar usuario
def eliminarUsuario(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM usuario WHERE id_usuario=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount
        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarUsuario : {e}")
        return []    

def eliminarArea(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM area WHERE id_area=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount
        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarArea : {e}")
        return []
    
def dataReportes():
    """
    Retorna los registros de accesos usando el nuevo esquema (tabla acceso y usuario).
    """
    try:
        with connectionBD() as conexion_MYSQLdb:
            with conexion_MYSQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                SELECT 
                    a.id_acceso, 
                    u.cedula, 
                    a.fecha_acceso AS fecha, 
                    a.tipo_acceso, 
                    r.nombre_rol AS rol
                FROM acceso a 
                LEFT JOIN usuario u ON u.id_usuario = a.id_usuario
                LEFT JOIN rol r ON u.id_rol = r.id_rol
                ORDER BY a.fecha_acceso DESC
                """
                cursor.execute(querySQL)
                reportes = cursor.fetchall()
        return reportes
    except Exception as e:
        print(f"Error en dataReportes : {e}")
        return []

def lastAccessBD(cedula):
    """
    Obtiene el último acceso de un usuario por cédula usando el nuevo esquema.
    """
    try:
        with connectionBD() as conexion_MYSQLdb:
            with conexion_MYSQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT a.id_acceso, u.cedula, a.fecha_acceso AS fecha FROM acceso a LEFT JOIN usuario u ON u.id_usuario = a.id_usuario WHERE u.cedula=%s ORDER BY a.fecha_acceso DESC LIMIT 1"
                cursor.execute(querySQL, (cedula,))
                reporte = cursor.fetchone()
        return reporte
    except Exception as e:
        print(f"Error en lastAccessBD : {e}")
        return []
import random
import string
def crearClave():
    caracteres = string.ascii_letters + string.digits  # Letras mayúsculas, minúsculas y dígitos
    longitud = 6  # Longitud de la clave

    clave = ''.join(random.choice(caracteres) for _ in range(longitud))
    print("La clave generada es:", clave)
    return clave
##GUARDAR CLAVES GENERADAS EN AUDITORIA
def guardarClaveAuditoria(clave_audi,id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                    # MODIFICACIÓN: Asegurar que tipo_acceso se inserte
                    sql = "INSERT INTO acceso (fecha_acceso, clave, id_usuario, tipo_acceso) VALUES (NOW(),%s,%s,'clave_generada')"
                    valores = (clave_audi,id)
                    mycursor.execute(sql, valores)
                    conexion_MySQLdb.commit()
                    resultado_insert = mycursor.rowcount
                    return resultado_insert 
        
    except Exception as e:
        return f'Se produjo un error en crear Clave: {str(e)}'
    
def lista_rolesBD():
    try:
        with connectionBD() as conexion_MYSQLdb:
            with conexion_MYSQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT * FROM rol"
                cursor.execute(querySQL)
                roles = cursor.fetchall()
                return roles
    except Exception as e:
        print(f"Error en select roles : {e}")
        return []

##CREAR AREA
def guardarArea(area_name):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                    sql = "INSERT INTO area (nombre_area) VALUES (%s)"
                    valores = (area_name,)
                    mycursor.execute(sql, valores)
                    conexion_MySQLdb.commit()
                    resultado_insert = mycursor.rowcount
                    return resultado_insert 
        
    except Exception as e:
        return f'Se produjo un error en crear Area: {str(e)}' 
    
##ACTUALIZAR AREA
def actualizarArea(area_id, area_name):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                sql = """UPDATE area SET nombre_area = %s WHERE id_area = %s"""
                valores = (area_name, area_id)
                mycursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                resultado_update = mycursor.rowcount
                return resultado_update 
        
    except Exception as e:
        return f'Se produjo un error al actualizar el área: {str(e)}'
    
    #--------consulta de datos de los roles-----------:

    
#--------------------- metodo de graficas ----------------------
def obtenerroles():
    try:
        with connectionBD() as conexion_MYSQLdb:
            with conexion_MYSQLdb.cursor(dictionary=True) as cursor:
                query = """
                    SELECT r.nombre_rol
                    FROM rol r
                    ORDER BY r.nombre_rol ASC
                """
                cursor.execute(query)
                roles = cursor.fetchall()
        return roles
    except Exception as e:
        print(f"Error en obtenerroles: {e}")
        return []
    
#------------------------ area de graficas -----------------------
def obtener_areas():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                query = """
                    SELECT nombre_area, numero_personas
                    FROM area
                    ORDER BY nombre_area ASC
                """
                cursor.execute(query)
                areas = cursor.fetchall()
        return areas
    except Exception as e:
        print(f"Error en obtener_areas: {e}")
        return []
    #------------------------ entrada de accesos --------------------------
def obtener_accesos_por_fecha(fecha_inicio, fecha_fin):
    """
    Obtiene el conteo de accesos por clave en un rango de fechas usando el nuevo esquema.
    """
    try:
        with connectionBD() as conexion_MYSQLdb:
            with conexion_MYSQLdb.cursor(dictionary=True) as cursor:
                query = """
                    SELECT clave, COUNT(id_acceso) AS cantidad
                    FROM acceso
                    WHERE fecha_acceso BETWEEN %s AND %s
                    GROUP BY clave
                    ORDER BY clave ASC
                """
                cursor.execute(query, (fecha_inicio, fecha_fin))
                accesos = cursor.fetchall()
        return accesos
    except Exception as e:
        print(f"Error en obtener_accesos_por_fecha: {e}")
        return []
    



# --- Funciones CRUD para Niños ---

def lista_ninoBD():
    """
    Lista todos los niños con los campos del nuevo esquema 'nino'.
    """
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                    SELECT 
                        id_nino, codigo_nino, nombre, apellido, fecha_nacimiento, 
                        id_genero, peso, altura, tutor_responsable, telefono_contacto, 
                        email_contacto, observaciones, activo, fecha_registro
                    FROM nino 
                    ORDER BY id_nino DESC
                """
                cursor.execute(querySQL)
                ninoBD = cursor.fetchall()
        return ninoBD
    except Exception as e:
        print(f"Error en lista_ninoBD: {e}")
        return []

def buscar_ninoBD(search_term):
    """
    Busca niños por nombre o apellido (nuevo esquema: nino).
    """
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = """
                    SELECT 
                        id_nino, codigo_nino, nombre, apellido, fecha_nacimiento, 
                        id_genero, peso, altura, tutor_responsable, telefono_contacto, 
                        email_contacto, observaciones, activo, fecha_registro
                    FROM nino
                    WHERE nombre LIKE %s OR apellido LIKE %s
                    ORDER BY id_nino DESC
                """
                search_pattern = f"%{search_term}%"
                mycursor.execute(querySQL, (search_pattern, search_pattern))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda
    except Exception as e:
        print(f"Ocurrió un error en buscar_ninoBD: {e}")
        return []

def obtener_nino_por_idBD(nino_id):
    """
    Obtiene un niño por su ID (nuevo esquema: nino).
    """
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                    SELECT 
                        id_nino, codigo_nino, nombre, apellido, fecha_nacimiento, 
                        id_genero, peso, altura, tutor_responsable, telefono_contacto, 
                        email_contacto, observaciones, activo, fecha_registro
                    FROM nino 
                    WHERE id_nino = %s
                """
                cursor.execute(querySQL, (nino_id,))
                nino = cursor.fetchone()
        return nino
    except Exception as e:
        print(f"Error en obtener_nino_por_idBD: {e}")
        return None

def guardar_ninoBD(
    codigo_nino, nombre, apellido, fecha_nacimiento, id_genero, peso, altura,
    tutor_responsable, telefono_contacto, email_contacto, observaciones, activo, fecha_registro=None
):
    """
    Agrega un nuevo niño con el esquema actualizado (tabla nino).
    """
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                sql = """
                    INSERT INTO nino (
                        codigo_nino, nombre, apellido, fecha_nacimiento, id_genero, peso, altura,
                        tutor_responsable, telefono_contacto, email_contacto, observaciones, activo, fecha_registro
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                valores = (
                    codigo_nino, nombre, apellido, fecha_nacimiento, id_genero, peso, altura,
                    tutor_responsable, telefono_contacto, email_contacto, observaciones, activo, fecha_registro
                )
                mycursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                resultado_insert = mycursor.rowcount
                return resultado_insert
    except Exception as e:
        print(f'Se produjo un error al crear el niño: {str(e)}')
        return 0

def actualizar_ninoBD(
    nino_id, nombre, apellido, fecha_nacimiento, id_genero, peso, altura,
    tutor_responsable, telefono_contacto, email_contacto, observaciones, activo
):
    """
    Actualiza los datos de un niño existente (tabla nino, nuevo esquema).
    """
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                sql = """
                    UPDATE nino SET 
                        nombre = %s, 
                        apellido = %s, 
                        fecha_nacimiento = %s, 
                        id_genero = %s, 
                        peso = %s, 
                        altura = %s, 
                        tutor_responsable = %s, 
                        telefono_contacto = %s, 
                        email_contacto = %s, 
                        observaciones = %s, 
                        activo = %s
                    WHERE id_nino = %s
                """
                valores = (
                    nombre, apellido, fecha_nacimiento, id_genero, peso, altura,
                    tutor_responsable, telefono_contacto, email_contacto, observaciones, activo, nino_id
                )
                mycursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                return mycursor.rowcount
    except Exception as e:
        print(f'Se produjo un error al actualizar el niño: {str(e)}')
        return 0

def eliminar_ninoBD(nino_id):
    """
    Elimina un niño por su ID (tabla nino, nuevo esquema).
    """
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM nino WHERE id_nino=%s"
                cursor.execute(querySQL, (nino_id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount
        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminar_ninoBD: {e}")
        return 0

def generarReporteExcel_nino():
    """
    Función para generar un reporte Excel de todos los niños con los nuevos atributos.
    """
    data_nino = lista_ninoBD()
    wb = openpyxl.Workbook()
    hoja = wb.active

    # Actualizar la fila de encabezado con los títulos de los nuevos campos
    cabeceraExcel = (
        "ID", "CÓDIGO", "NOMBRE", "APELLIDO", "FECHA NACIMIENTO", "ID GÉNERO", 
        "PESO (kg)", "ALTURA (cm)", "OBSERVACIONES GENERALES", "TUTOR RESPONSABLE", 
        "TELÉFONO CONTACTO", "EMAIL CONTACTO", "ACTIVO", "FECHA REGISTRO", "ÚLTIMA ACTUALIZACIÓN"
    )
    hoja.append(cabeceraExcel)

    # Agregar los registros a la hoja, mapeando los nuevos campos
    for nino in data_nino:
        id_nino = nino.get('id_nino')
        codigo_nino = nino.get('codigo_nino')
        nombre = nino.get('nombre')
        apellido = nino.get('apellido')
        fecha_nacimiento = nino.get('fecha_nacimiento')
        id_genero = nino.get('id_genero')
        peso = nino.get('peso')
        altura = nino.get('altura')
        observaciones_generales = nino.get('observaciones_generales')
        tutor_responsable = nino.get('tutor_responsable')
        telefono_contacto = nino.get('telefono_contacto')
        email_contacto = nino.get('email_contacto')
        activo = "Sí" if nino.get('activo') else "No" # Convertir booleano a texto
        fecha_registro = nino.get('fecha_registro')
        ultima_actualizacion = nino.get('ultima_actualizacion')

        # Agregar los valores a la hoja
        hoja.append((
            id_nino, codigo_nino, nombre, apellido, fecha_nacimiento, id_genero, 
            peso, altura, observaciones_generales, tutor_responsable, 
            telefono_contacto, email_contacto, activo, fecha_registro, ultima_actualizacion
        ))

    fecha_actual = datetime.datetime.now()
    archivoExcel = f"Reporte_nino_{fecha_actual.strftime('%Y_%m_%d')}.xlsx"
    carpeta_descarga = "static/downloads-excel" # Ajusta según tu estructura de carpetas
    ruta_descarga = os.path.join(os.path.dirname(os.path.abspath(__file__)), carpeta_descarga)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
        os.chmod(ruta_descarga, 0o755)

    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    # Enviar el archivo como respuesta HTTP
    return send_file(ruta_archivo, as_attachment=True)

# --- Funciones CRUD para Condiciones ---

def lista_condicionBD():
    """
    Lista todas las condicion médicas/cognitivas (nuevo esquema: condicion).
    """
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                    SELECT 
                        id_condicion, nombre_condicion, descripcion, categoria
                    FROM condicion 
                    ORDER BY id_condicion DESC
                """
                cursor.execute(querySQL)
                condicionBD = cursor.fetchall()
        return condicionBD
    except Exception as e:
        print(f"Error en lista_condicionBD: {e}")
        return []

def buscar_condicionBD(search_term):
    """
    Busca condicion por nombre, descripción o categoría (nuevo esquema: condicion).
    """
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = """
                    SELECT 
                        id_condicion, nombre_condicion, descripcion, categoria
                    FROM condicion
                    WHERE nombre_condicion LIKE %s OR descripcion LIKE %s OR categoria LIKE %s
                    ORDER BY id_condicion DESC
                """
                search_pattern = f"%{search_term}%"
                mycursor.execute(querySQL, (search_pattern, search_pattern, search_pattern))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda
    except Exception as e:
        print(f"Ocurrió un error en buscar_condicionBD: {e}")
        return []

def obtener_condicion_por_idBD(condicion_id):
    """
    Obtiene una condición por su ID (nuevo esquema: condicion).
    """
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                    SELECT 
                        id_condicion, nombre_condicion, descripcion, categoria
                    FROM condicion 
                    WHERE id_condicion = %s
                """
                cursor.execute(querySQL, (condicion_id,))
                condicion = cursor.fetchone()
        return condicion
    except Exception as e:
        print(f"Error en obtener_condicion_por_idBD: {e}")
        return None

def guardar_condicionBD(nombre_condicion, descripcion, categoria):
    """
    Función para agregar una nueva condición con todos sus atributos.
    """
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                sql = """
                    INSERT INTO condicion (
                        nombre_condicion, descripcion, categoria
                    ) VALUES (%s, %s, %s)
                """
                valores = (nombre_condicion, descripcion, categoria)
                mycursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                resultado_insert = mycursor.rowcount
                return resultado_insert
    except Exception as e:
        print(f'Se produjo un error al crear la condición: {str(e)}')
        return 0

def actualizar_condicionBD(condicion_id, nombre_condicion, descripcion, categoria):
    """
    Función para actualizar los datos de una condición existente, incluyendo todos sus atributos.
    """
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                sql = """
                    UPDATE condicion SET 
                        nombre_condicion = %s, 
                        descripcion = %s, 
                        categoria = %s
                    WHERE id_condicion = %s
                """
                valores = (nombre_condicion, descripcion, categoria, condicion_id)
                mycursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                resultado_update = mycursor.rowcount
                return resultado_update
    except Exception as e:
        print(f'Se produjo un error al actualizar la condición: {str(e)}')
        return 0

def eliminar_condicionBD(condicion_id):
    """
    Función para eliminar una condición por su ID.
    """
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM condicion WHERE id_condicion=%s"
                cursor.execute(querySQL, (condicion_id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount
        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminar_condicionBD: {e}")
        return 0

def generarReporteExcel_condicion():
    """
    Función para generar un reporte Excel de todas las condicion.
    """
    data_condicion = lista_condicionBD()
    wb = openpyxl.Workbook()
    hoja = wb.active

    # Agregar la fila de encabezado con los títulos de los campos
    cabeceraExcel = ("ID", "NOMBRE CONDICIÓN", "DESCRIPCIÓN", "CATEGORÍA")
    hoja.append(cabeceraExcel)

    # Agregar los registros a la hoja, mapeando los campos
    for condicion in data_condicion:
        id_condicion = condicion.get('id_condicion')
        nombre_condicion = condicion.get('nombre_condicion')
        descripcion = condicion.get('descripcion')
        categoria = condicion.get('categoria')

        # Agregar los valores a la hoja
        hoja.append((id_condicion, nombre_condicion, descripcion, categoria))

    fecha_actual = datetime.datetime.now()
    archivoExcel = f"Reporte_condicion_{fecha_actual.strftime('%Y_%m_%d')}.xlsx"
    carpeta_descarga = "static/downloads-excel" # Ajusta según tu estructura de carpetas
    ruta_descarga = os.path.join(os.path.dirname(os.path.abspath(__file__)), carpeta_descarga)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
        os.chmod(ruta_descarga, 0o755)

    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    # Enviar el archivo como respuesta HTTP
    return send_file(ruta_archivo, as_attachment=True)

# CONDICIONES NIÑOS
def lista_nino_condicionBD():
    conexion = connectionBD()
    cursor = conexion.cursor(dictionary=True)
    sql = """
        SELECT 
            nc.id_nino_condicion,
            n.nombre AS id_nino,
            c.nombre_condicion AS id_condicion, # Cambiado a nombre_condicion para mostrar el nombre
            nc.severidad,
            nc.fecha_diagnostico,
            nc.observaciones,
            nc.activo
        FROM nino_condicion nc
        JOIN nino n ON nc.id_nino = n.id_nino
        JOIN condicion c ON nc.id_condicion = c.id_condicion
    """
    cursor.execute(sql)
    resultado = cursor.fetchall()
    conexion.close()
    return resultado

def insertar_nino_condicionBD(datos):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                sql = """
                    INSERT INTO nino_condicion (
                        id_nino, id_condicion, severidad, fecha_diagnostico, observaciones, activo
                    ) VALUES (%s, %s, %s, %s, %s, %s)
                """
                valores = (
                    datos["id_nino"], datos["id_condicion"], datos["severidad"], 
                    datos["fecha_diagnostico"], datos["observaciones"], datos["activo"]
                )
                mycursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                return mycursor.rowcount
    except Exception as e:
        print(f'Se produjo un error al insertar la condición del niño: {str(e)}')
        return 0

def obtener_nino_condicionBD(id_nino_condicion):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                    SELECT 
                        nc.id_nino_condicion,
                        n.nombre AS nombre_nino,
                        c.nombre_condicion AS nombre_condicion,
                        nc.severidad,
                        nc.fecha_diagnostico,
                        nc.observaciones,
                        nc.activo,
                        nc.id_nino,
                        nc.id_condicion
                    FROM nino_condicion nc
                    JOIN nino n ON nc.id_nino = n.id_nino
                    JOIN condicion c ON nc.id_condicion = c.id_condicion
                    WHERE nc.id_nino_condicion = %s
                """
                cursor.execute(querySQL, (id_nino_condicion,))
                nino_cond = cursor.fetchone()
        return nino_cond
    except Exception as e:
        print(f"Error en obtener_nino_condicionBD: {e}")
        return None

def actualizar_nino_condicionBD(id_nino_condicion, datos):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                sql = """
                    UPDATE nino_condicion SET 
                        severidad = %s, 
                        fecha_diagnostico = %s, 
                        observaciones = %s, 
                        activo = %s
                    WHERE id_nino_condicion = %s
                """
                valores = (
                    datos["severidad"], datos["fecha_diagnostico"], 
                    datos["observaciones"], datos["activo"], id_nino_condicion
                )
                mycursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                return mycursor.rowcount
    except Exception as e:
        print(f'Se produjo un error al actualizar la condición del niño: {str(e)}')
        return 0

def eliminar_nino_condicionBD(id_nino_condicion):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM nino_condicion WHERE id_nino_condicion=%s"
                cursor.execute(querySQL, (id_nino_condicion,))
                conexion_MySQLdb.commit()
                return cursor.rowcount
    except Exception as e:
        print(f"Error en eliminar_nino_condicionBD: {e}")
        return 0

def obtener_nino_condicionBD(id_nino_condicion):
    """
    Obtiene una relación niño-condición por su ID.
    """
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                sql = """
                    SELECT nc.id_nino_condicion, nc.id_nino, nc.id_condicion, c.nombre_condicion, c.descripcion, c.categoria
                    FROM nino_condicion nc
                    JOIN condicion c ON nc.id_condicion = c.id_condicion
                    WHERE nc.id_nino_condicion = %s
                """
                cursor.execute(sql, (id_nino_condicion,))
                return cursor.fetchone()
    except Exception as e:
        print(f"Error en obtener_nino_condicionBD: {e}")
        return None

def lista_nino_condicionBD():
    """
    Función para listar todas las relaciones niño-condición.
    Incluye nombres de niño y condición para una mejor visualización.
    """
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                sql = """
                    SELECT 
                        nc.id_nino_condicion,
                        n.nombre AS nombre_nino,
                        n.apellido AS apellido_nino,
                        c.nombre_condicion AS nombre_condicion,
                        nc.severidad,
                        nc.fecha_diagnostico,
                        nc.observaciones,
                        nc.activo
                    FROM nino_condicion nc
                    JOIN nino n ON nc.id_nino = n.id_nino
                    JOIN condicion c ON nc.id_condicion = c.id_condicion
                    ORDER BY nc.fecha_diagnostico DESC
                """
                cursor.execute(sql)
                resultado = cursor.fetchall()
                # print(f"DEBUG: Datos de nino_condicion obtenidos: {resultado}") # Mantener o quitar según necesidad
        return resultado
    except Exception as e:
        print(f"Error en lista_nino_condicionBD: {e}")
        return []

def insertar_nino_condicionBD(datos):
    """
    Inserta una nueva relación niño-condición en la base de datos.
    """
    # Definir los valores permitidos para el ENUM de severidad
    severidades_validas = ['leve', 'moderada', 'severa']
    
    # Normalizar la severidad: a minúsculas y sin espacios extra
    severidad_normalizada = datos["severidad"].strip().lower()

    if severidad_normalizada not in severidades_validas:
        print(f"DEBUG: Valor de severidad inválido: '{datos['severidad']}'. Valores permitidos: {severidades_validas}")
        return 0 # Indica fallo debido a valor inválido

    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                sql = """
                    INSERT INTO nino_condicion (
                        id_nino, id_condicion, severidad, fecha_diagnostico, observaciones, activo
                    ) VALUES (%s, %s, %s, %s, %s, %s)
                """
                valores = (
                    datos["id_nino"], datos["id_condicion"], severidad_normalizada, # Usar el valor normalizado
                    datos["fecha_diagnostico"], datos["observaciones"], datos["activo"]
                )
                mycursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                print(f"DEBUG: Filas afectadas por insertar_nino_condicionBD: {mycursor.rowcount}") # DEBUG
                return mycursor.rowcount
    except Exception as e:
        print(f'DEBUG: Se produjo un error al insertar la condición del niño: {str(e)}') # DEBUG
        return 0

def obtener_nino_condicionBD(id_nino_condicion):
    """
    Obtiene una relación niño-condición por su ID.
    Incluye los IDs de niño y condición para preselección en formularios.
    """
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                    SELECT 
                        nc.id_nino_condicion,
                        nc.id_nino,
                        n.nombre AS nombre_nino,
                        n.apellido AS apellido_nino,
                        nc.id_condicion,
                        c.nombre_condicion AS nombre_condicion,
                        nc.severidad,
                        nc.fecha_diagnostico,
                        nc.observaciones,
                        nc.activo
                    FROM nino_condicion nc
                    JOIN nino n ON nc.id_nino = n.id_nino
                    JOIN condicion c ON nc.id_condicion = c.id_condicion
                    WHERE nc.id_nino_condicion = %s
                """
                cursor.execute(querySQL, (id_nino_condicion,))
                nino_cond = cursor.fetchone()
        return nino_cond
    except Exception as e:
        print(f"Error en obtener_nino_condicionBD: {e}")
        return None

def actualizar_nino_condicionBD(id_nino_condicion, datos):
    """
    Actualiza una relación niño-condición existente.
    """
    # Definir los valores permitidos para el ENUM de severidad
    severidades_validas = ['leve', 'moderada', 'severa']
    
    # Normalizar la severidad: a minúsculas y sin espacios extra
    severidad_normalizada = datos["severidad"].strip().lower()

    if severidad_normalizada not in severidades_validas:
        print(f"DEBUG: Valor de severidad inválido para actualización: '{datos['severidad']}'. Valores permitidos: {severidades_validas}")
        return 0 # Indica fallo debido a valor inválido

    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                sql = """
                    UPDATE nino_condicion SET 
                        severidad = %s, 
                        fecha_diagnostico = %s, 
                        observaciones = %s, 
                        activo = %s
                    WHERE id_nino_condicion = %s
                """
                valores = (
                    severidad_normalizada, # Usar el valor normalizado
                    datos["fecha_diagnostico"], 
                    datos["observaciones"], 
                    datos["activo"], 
                    id_nino_condicion
                )
                mycursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                return mycursor.rowcount
    except Exception as e:
        print(f'Se produjo un error al actualizar la condición del niño: {str(e)}')
        return 0

def eliminar_nino_condicionBD(id_nino_condicion):
    """
    Elimina una relación niño-condición de la base de datos.
    """
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM nino_condicion WHERE id_nino_condicion=%s"
                cursor.execute(querySQL, (id_nino_condicion,))
                conexion_MySQLdb.commit()
                return cursor.rowcount
    except Exception as e:
        print(f"Error en eliminar_nino_condicionBD: {e}")
        return 0


# === FUNCIONES DE PROCESAMIENTO Y GRAFICACIÓN PARA FIREBASE REPORTES ===
def procesar_porcentaje(diccionario):
    total = 0
    valores = {}
    for k, v in diccionario.items():
        if isinstance(v, dict):
            valor = v.get('count', 0)
        elif isinstance(v, (int, float)):
            valor = v
        else:
            valor = 0
        valores[k] = valor
        total += valor
    return {k: round((val / total) * 100, 2) if total > 0 else 0 for k, val in valores.items()}

def graficar_porcentajes(porcentajes, titulo, color, nombre_archivo_path):
    fig, ax = plt.subplots()
    items = list(porcentajes.keys())
    valores = list(porcentajes.values())
    ax.bar(items, valores, color=color)
    ax.set_title(titulo)
    ax.set_ylabel('Porcentaje (%)')
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.savefig(nombre_archivo_path)
    plt.close()

# === FUNCIONES PARA OBTENER SESIONES Y FECHAS DE FIREBASE ===
def get_session_types():
    try:
        ref = db.reference("sesiones")
        data = ref.get()
        return sorted(data.keys()) if data else []
    except Exception as e:
        logger.error(f"Error al obtener tipos de sesión de Firebase: {e}")
        return []

def get_session_dates(session_name):
    try:
        ref = db.reference(f"sesiones/{session_name}")
        data = ref.get()
        return sorted(data.keys()) if data else []
    except Exception as e:
        logger.error(f"Error al obtener fechas de sesión de Firebase para '{session_name}': {e}")
        return []

# === Obtener modos disponibles para una fecha específica en Firebase ===
def get_modes_for_date(session_date):
    """
    Dada una fecha, retorna una lista de modos (tipos de sesión) que tienen datos para esa fecha.
    Además, imprime logs detallados para depuración.
    """
    try:
        ref = db.reference("sesiones")
        data = ref.get()
        modos_con_fecha = []
        print(f"[DEBUG] Buscando modos para la fecha: {session_date}")
        if data:
            for modo, fechas in data.items():
                print(f"[DEBUG] Modo: {modo} - Fechas disponibles: {list(fechas.keys()) if isinstance(fechas, dict) else fechas}")
                if isinstance(fechas, dict) and session_date in fechas:
                    modos_con_fecha.append(modo)
        print(f"[DEBUG] Modos encontrados para la fecha {session_date}: {modos_con_fecha}")
        return sorted(modos_con_fecha)
    except Exception as e:
        logger.error(f"Error al obtener modos para la fecha '{session_date}' en Firebase: {e}")
        return []

# === FUNCIÓN PRINCIPAL DE GENERACIÓN DE REPORTE DE FIREBASE ===
def contar_activaciones_estado_0(datos_sensores):
    """
    Calcula las activaciones de dos maneras:
    1. Suma 'activaciones_acumuladas' si existe (nueva estructura).
    2. Cuenta los hijos del nodo 'activaciones' si existe (antigua estructura).
    """
    conteo = {}
    if not isinstance(datos_sensores, dict):
        return conteo

    for sensor, sensor_data in datos_sensores.items():
        total_activaciones = 0
        is_new_format = False

        if isinstance(sensor_data, dict):
            # Primero, determinar el formato de los datos para este sensor
            # Buscamos la presencia de 'activaciones_acumuladas' en cualquier sub-registro
            for push_id_data in sensor_data.values():
                if isinstance(push_id_data, dict) and "activaciones_acumuladas" in push_id_data:
                    is_new_format = True
                    break  # Formato nuevo detectado, no es necesario seguir buscando

            # Ahora, procesar según el formato detectado
            if is_new_format:
                # Si es formato nuevo, sumar todas las 'activaciones_acumuladas'
                for push_id_data in sensor_data.values():
                    if isinstance(push_id_data, dict):
                        total_activaciones += push_id_data.get("activaciones_acumuladas", 0)
            else:
                # Si no, buscar el formato antiguo (nodo 'activaciones')
                activaciones_node = sensor_data.get("activaciones")
                if isinstance(activaciones_node, dict):
                    # Contar el número de hijos (push IDs)
                    total_activaciones = len(activaciones_node)

        conteo[sensor] = total_activaciones
        
    return conteo

def generar_reporte_firebase(session_type, session_date):
    """
    Genera el reporte de Firebase para un tipo de sesión y fecha específicos,
    basado en la lógica del script original de Tkinter.
    Devuelve un objeto BytesIO que contiene el archivo Excel y el nombre del archivo.
    """
    try:
        ref = db.reference(f"sesiones/{session_type}/{session_date}")
        datos = ref.get()

        if not datos:
            raise ValueError(f"No se encontraron datos para la sesión '{session_type}' en la fecha '{session_date}'.")

        # Extraer y contar las activaciones correctamente
        botones_data = datos.get("botones", {})
        escalones_data = datos.get("escalones", {})

        conteo_botones = contar_activaciones_estado_0(botones_data)
        conteo_escalones = contar_activaciones_estado_0(escalones_data)

        porcentaje_botones = procesar_porcentaje(conteo_botones)
        porcentaje_escalones = procesar_porcentaje(conteo_escalones)

        charts_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'static', 'temp_charts')
        os.makedirs(charts_dir, exist_ok=True)

        grafico_botones_filename = f'grafico_botones_{session_type}_{session_date}_{uuid.uuid4().hex}.png'.replace(":", "-")
        grafico_escalones_filename = f'grafico_escalones_{session_type}_{session_date}_{uuid.uuid4().hex}.png'.replace(":", "-")

        grafico_botones_path = os.path.join(charts_dir, grafico_botones_filename)
        grafico_escalones_path = os.path.join(charts_dir, grafico_escalones_filename)

        graficar_porcentajes(porcentaje_botones, 'Porcentaje de uso de botones', 'skyblue', grafico_botones_path)
        graficar_porcentajes(porcentaje_escalones, 'Porcentaje de uso de escalones', 'lightgreen', grafico_escalones_path)

        formato_excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'downloads-excel', 'Formato_Excel_General.xlsx')
        
        if not os.path.exists(formato_excel_path):
            logger.error(f"El archivo de formato Excel no se encontró en: {formato_excel_path}")
            raise FileNotFoundError(f"El archivo de formato Excel 'Formato_Excel_General.xlsx' no se encontró en la ubicación esperada.")

        wb = load_workbook(formato_excel_path)
        ws = wb.active

        ws["C3"] = session_type.capitalize()
        ws["C4"] = session_date

        for i in range(1, 7):
            fila = 10 + i
            key = f"boton{i}"
            ws[f"C{fila}"] = conteo_botones.get(key, 0)
            ws[f"D{fila}"] = porcentaje_botones.get(key, 0)

        for i in range(1, 5):
            fila = 22 + i
            key = f"escalon{i}"
            ws[f"C{fila}"] = conteo_escalones.get(key, 0)
            ws[f"D{fila}"] = porcentaje_escalones.get(key, 0)

        ws['C29'] = datos.get('fin_sesion', 'No definido')
        ws['C30'] = datos.get('inicio_sesion', 'No definido')

        if "Gráficos" in wb.sheetnames:
            ws_graf = wb["Gráficos"]
        else:
            ws_graf = wb.create_sheet("Gráficos")

        for drawing in list(ws_graf._images):
            ws_graf._images.remove(drawing)

        img_botones = ExcelImage(grafico_botones_path)
        ws_graf.add_image(img_botones, 'A1')

        img_escalones = ExcelImage(grafico_escalones_path)
        ws_graf.add_image(img_escalones, 'A20')

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        try:
            os.remove(grafico_botones_path)
            os.remove(grafico_escalones_path)
        except OSError as e:
            logger.warning(f"Error al eliminar archivos temporales de gráficos: {e}")

        # Devolver el objeto BytesIO y el nombre del archivo sugerido
        file_name = f"REPORTE_{session_type}_{session_date}.xlsx".replace(":", "-")
        return output, file_name

    except FileNotFoundError as e:
        logger.error(f"Error de archivo en generar_reporte_firebase: {e}")
        raise # Re-lanza la excepción para que el router la capture
    except ValueError as e:
        logger.error(f"Error de datos de Firebase en generar_reporte_firebase: {e}")
        raise # Re-lanza la excepción
    except Exception as e:
        logger.exception(f"Error inesperado al generar el reporte de Firebase: {e}") # Usa exception para el traceback completo
        raise # Re-lanza la excepción
#Hasta aqui
def generar_codigo_nino():
    """
    Genera un código único para el niño en el formato NINO-YYYYMMDD-XYZ
    """
    try:
        # Usar hora de Ecuador para el código (evita desfase horario en Cloud Run)
        from app import get_ecuador_time
        fecha_actual = get_ecuador_time().strftime('%Y%m%d')

        with connectionBD() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT COUNT(*) FROM nino WHERE DATE(fecha_registro) = CURDATE()")
                total_hoy = cursor.fetchone()[0] + 1  # +1 para el nuevo niño

        secuencia = str(total_hoy).zfill(3)  # siempre 3 dígitos
        codigo = f"NINO-{fecha_actual}-{secuencia}"
        return codigo

    except Exception as e:
        print(f"Error generando código de niño: {e}")
        # En caso de error, usar un fallback
        # Fallback también usa hora de Ecuador
        from app import get_ecuador_time
        return f"NINO-{get_ecuador_time().strftime('%Y%m%d%H%M%S')}"


def buscar_usuariosBD(termino_busqueda):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                query = """
                    SELECT 
                        u.id_usuario,
                        u.cedula,
                        u.nombre,
                        u.apellido,
                        u.email,
                        u.telefono,
                        u.fecha_creacion,
                        u.activo,
                        u.id_rol,
                        u.especialidad,
                        u.anos_experiencia,
                        u.certificaciones
                    FROM usuario u
                    WHERE 
                        LOWER(u.cedula) LIKE %s OR 
                        LOWER(u.nombre) LIKE %s OR 
                        LOWER(u.apellido) LIKE %s OR 
                        LOWER(u.email) LIKE %s
                """
                like_term = f"%{termino_busqueda.lower()}%"
                mycursor.execute(query, (like_term, like_term, like_term, like_term))
                return mycursor.fetchall()
    except Exception as e:
        print(f"Error al buscar usuarios: {str(e)}")
        return []
    
# Se agregaron estas líneas para obtener datos de gráficos
def get_chart_data_for_nino(nino_id):
    try:
        # 1. Obtener el código del niño desde la base de datos local
        nino_info = obtener_nino_por_idBD(nino_id)
        if not nino_info or 'codigo_nino' not in nino_info:
            raise ValueError(f"No se encontró información o el código para el niño con ID {nino_id}")
        codigo_nino = nino_info['codigo_nino']
        nombre_nino = f"{nino_info['nombre']} {nino_info['apellido']}"

        # 2. Obtener las sesiones del niño desde Firebase
        sesiones_ref = db.reference(f'nino/{codigo_nino}/sesiones')
        sesiones_ids = sesiones_ref.get()

        if not sesiones_ids:
            return {"nombre_nino": nombre_nino, "total_activaciones": 0, "botones_data": None, "escalones_data": None}

        # 3. Recopilar y agregar datos de todas las sesiones
        total_botones = {}
        total_escalones = {}

        # Buscar en ambos modos, juego y evaluación
        modos_sesion = ['modojuego', 'modoevaluacion']
        for sesion_id in sesiones_ids:
            for modo in modos_sesion:
                sesion_ref = db.reference(f'sesiones/{modo}/{sesion_id}')
                sesion_data = sesion_ref.get()

                if sesion_data:
                    # Agregar datos de botones
                    if 'botones' in sesion_data:
                        for boton, data in sesion_data['botones'].items():
                            if 'conteo' in data:
                                total_botones[boton] = total_botones.get(boton, 0) + data['conteo']
                    
                    # Agregar datos de escalones
                    if 'escalones' in sesion_data:
                        for escalon, data in sesion_data['escalones'].items():
                            if 'conteo' in data:
                                total_escalones[escalon] = total_escalones.get(escalon, 0) + data['conteo']

        # 4. Procesar los datos agregados
        porcentaje_botones = procesar_porcentaje(total_botones)
        porcentaje_escalones = procesar_porcentaje(total_escalones)
        total_activaciones = sum(total_botones.values()) + sum(total_escalones.values())

        return {
            "nombre_nino": nombre_nino,
            "botones_data": {
                "counts": total_botones,
                "percentages": porcentaje_botones
            } if total_botones else None,
            "escalones_data": {
                "counts": total_escalones,
                "percentages": porcentaje_escalones
            } if total_escalones else None,
            "total_activaciones": total_activaciones
        }

    except Exception as e:
        logger.error(f"Error al obtener datos de gráficos para el niño {nino_id}: {e}")
        raise

def get_firebase_chart_data(session_type, session_date):
    """
    Obtiene los datos de Firebase para un tipo de sesión y fecha específicos,
    y devuelve los recuentos y porcentajes para botones y escalones,
    además de los totales de activación.
    """
    try:
        ref = db.reference(f"sesiones/{session_type}/{session_date}")
        datos = ref.get()

        if not datos:
            raise ValueError(f"No se encontraron datos para la sesión '{session_type}' en la fecha '{session_date}'.")

        # Normalizar claves a minúsculas para evitar problemas de mayúsculas/minúsculas
        botones_data = {k.lower(): v for k, v in datos.get("botones", {}).items()}
        escalones_data = {k.lower(): v for k, v in datos.get("escalones", {}).items()}

        # Usar la función de conteo para obtener las activaciones correctas
        conteo_botones = contar_activaciones_estado_0(botones_data)
        conteo_escalones = contar_activaciones_estado_0(escalones_data)

        # Calcular porcentajes basados en los conteos correctos
        porcentaje_botones = procesar_porcentaje(conteo_botones)
        porcentaje_escalones = procesar_porcentaje(conteo_escalones)

        # Calcular totales de activación
        total_activaciones = sum(conteo_botones.values()) + sum(conteo_escalones.values())

        return {
            "botones_data": {
                "counts": conteo_botones,
                "percentages": porcentaje_botones
            },
            "escalones_data": {
                "counts": conteo_escalones,
                "percentages": porcentaje_escalones
            },
            "total_activaciones": total_activaciones
        }

    except Exception as e:
        logger.error(f"Error al obtener datos de gráficos de Firebase: {e}")
        raise # Re-lanza la excepción para que el router la capture